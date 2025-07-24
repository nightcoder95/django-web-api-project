import csv
import openpyxl
from django.http import HttpResponse
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from .models import Product, Category

class BaseExportView(LoginRequiredMixin, View):
    """Base class for export views with permission checking"""
    
    def dispatch(self, request, *args, **kwargs):
        # Only admin and staff can export data
        if not request.user.role in ['admin', 'staff']:
            raise PermissionDenied("Only admin and staff users can export data.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_products_queryset(self):
        """Get products based on user role"""
        user = self.request.user
        if user.role == 'admin':
            return Product.objects.select_related('category', 'created_by').all()
        elif user.role == 'staff':
            return Product.objects.select_related('category', 'created_by').filter(status='uploaded')
        return Product.objects.none()


class ExportProductsCSVView(BaseExportView):
    """Export products to CSV format"""
    
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        filename = f'products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'ID',
            'Title',
            'Description',
            'Category',
            'Price',
            'Status',
            'Created By',
            'Created At',
            'Updated At'
        ])
        
        # Write product data
        products = self.get_products_queryset()
        for product in products:
            writer.writerow([
                product.id,
                product.title,
                product.description,
                product.category.name,
                product.price,
                product.get_status_display(),
                product.created_by.username,
                product.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                product.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response


class ExportProductsExcelView(BaseExportView):
    """Export products to Excel format"""
    
    def get(self, request):
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Products'
        
        # Define headers
        headers = [
            'ID', 'Title', 'Description', 'Category', 'Price', 
            'Status', 'Created By', 'Created At', 'Updated At'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Write product data
        products = self.get_products_queryset()
        for row, product in enumerate(products, 2):
            worksheet.cell(row=row, column=1, value=product.id)
            worksheet.cell(row=row, column=2, value=product.title)
            worksheet.cell(row=row, column=3, value=product.description)
            worksheet.cell(row=row, column=4, value=product.category.name)
            worksheet.cell(row=row, column=5, value=float(product.price))
            worksheet.cell(row=row, column=6, value=product.get_status_display())
            worksheet.cell(row=row, column=7, value=product.created_by.username)
            worksheet.cell(row=row, column=8, value=product.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            worksheet.cell(row=row, column=9, value=product.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response


class ExportCategoriesCSVView(BaseExportView):
    """Export categories to CSV format"""
    
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        filename = f'categories_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow(['ID', 'Name', 'Product Count'])
        
        # Write category data
        categories = Category.objects.all()
        for category in categories:
            product_count = category.product_set.count()
            writer.writerow([
                category.id,
                category.name,
                product_count
            ])
        
        return response


class ExportView(BaseExportView):
    """Main export page with download options"""
    template_name = 'products/export.html'
    
    def get(self, request):
        products_count = self.get_products_queryset().count()
        categories_count = Category.objects.count()
        
        context = {
            'products_count': products_count,
            'categories_count': categories_count,
        }
        return self.render_to_response(context)
    
    def render_to_response(self, context):
        from django.shortcuts import render
        return render(self.request, self.template_name, context)
